import requests
import json
from openpyxl import Workbook, load_workbook


def request_login():
    headers_login = {
        'accept': '*/*',
        'accept-language': 'ru',
        'content-type': 'application/json',
        'origin': 'https://attend.gdcevents.gdconf.com',
        'referer': 'https://attend.gdcevents.gdconf.com/login/ru-RU/login',
        'sec-ch-ua': '"Google Chrome";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
        'x-called-from-server': 'false',
    }

    json_data_login = {
        'emailAddress': 'michael@28software.com',
        'password': '1234Adimptzu',
        'shouldReturnRefreshToken': True,
    }

    response = requests.post('https://attend.gdcevents.gdconf.com/auth/api/login',
                             headers=headers_login, json=json_data_login)
    json_data = json.dumps(response.json(), indent=2)

    # Запись JSON-данных в файл
    dictionary = eval(json_data)
    auth_key = "Bearer "+dictionary['accessToken']
    return auth_key


def request_all(auth_key, search=None, endCursor=None,):
    headers_all = {
        'accept': '*/*',
        'accept-language': 'ru,en;q=0.9',
        'authorization': auth_key,
        'content-type': 'application/json',
        # 'cookie': 'intercom-id-ineyz1by=e35f5b9f-b066-457a-9aff-0a53a0f551fb; intercom-device-id-ineyz1by=e53daf4f-460c-489c-ad59-9c84706e7f2f; swapcard-cookie-consent=%7B%22accepted%22%3Atrue%7D; __stripe_mid=e1417346-a801-431d-8686-ab32cbfd508a693eca; next-i18next=ru-RU; _ga=GA1.1.746820343.1711986307; swapcard-auth-api-refresh-token=eyJhbGciOiJSUzUxMiIsInR5cCI6IkpXVCJ9.eyJzZXNzaW9uSWQiOiI2NjBhZTI1YzAwMmZhNzBmNDI3MTFmOTciLCJ0eXBlIjoicmVmcmVzaC10b2tlbiIsInVzZXJJZCI6IjY0ZDYzZWMxMjZlMDlmNmE1NDI0ZGJmYiIsImlhdCI6MTcxMTk4OTM0MCwiaXNzIjoiYXV0aC1hcGkifQ.OEUEI9HJDC5FEa680wFow4abrnoVhfOnx57j7J-eMW7Iw8LQYOQB6Bqfew0ocPtaYR_AihztAG2cVYXl8IVbYkeqlMsf4DfkZpRe7Y1D5TRvtMqs-RkkTB6g-o0lAABwMZS5BA3M6pJ7ZjX9ol0F-sGSz6sgjBm-tq64sYGHUMBkg4v0KbhvZgQYRtGbtqC6CWmw5r043-pZ7jBTxrY_dNGjLSq03spnPisjgnqqH0-icib7XiiuNScqsemJhwVwI4uAdAMmHZCXk-7Y2B4_AKcSJOjp6YMiO-K1CkU-XQ1GeSQoCPjEbJGV1QbKqwQsugQvPUv2PLCumIx6OjGlOfJHhr383w2wlUNn7x-7GXvfNFR1u28I6ru0jJhmhYHcCsMbp6dtI6i-nUx8XRJv5Wz_EI8J9GzRcJ6Rkw1Y0uN9O89GDhzItQ3Jk8STEqCpFitwzOOvWDctrjnbXRsDVHj71OOxLq8VzLKVufO_U0ORvvD8zz3d3PayHLfLoe2NZEtxnRBJeJVjezDYHHDSHykMbml2J71WgCPRA52MRL9QGkwWrKMsQpWKyoOu9Nsj7IAVIelAJRLu7xrMBf04l9TH8edHRrHRJdPJprFt2N8qnNcjZbKlaLU-kQqEqjNSqzl8LLOg2xvvfU1y-hYQuN5BB2dwUB_cnKV4tdUqQYA; __stripe_sid=76346eba-8ea1-4bc0-bef5-621d4edc15437a7e56; intercom-session-ineyz1by=eDJXb1N6M2owblM2YmgrdG5aeWxnWTJvY1Vza2RRcEFBWTk3aWxsbS81end6S1VwUUZGNFZzWkpFR3p4b2huby0tOWdUa1JmL3VvQVpxN1B0MldxQXRPdz09--a73ea1e3e8dfff6839f8fce3ffc7fb77ab63cc72; _ga_VWQ0TSBSYY=GS1.1.1712840847.12.1.1712842860.60.0.0',
        'origin': 'https://attend.gdcevents.gdconf.com',
        'referer': 'https://attend.gdcevents.gdconf.com/',
        'sec-ch-ua': '"Google Chrome";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
        'x-client-origin': 'attend.gdcevents.gdconf.com',
        'x-client-platform': 'Event App',
        'x-client-version': '2.308.37',
    }
    json_data_all = [

        {
            'operationName': 'EventPeopleListViewConnectionQuery',
            'variables': {
                'viewId': 'RXZlbnRWaWV3XzY2OTQyNw==',
                'search': search,
                'sort': None,
                'endCursor': endCursor,
            },
            'extensions': {
                'persistedQuery': {
                    'version': 1,
                    'sha256Hash': 'a4e69f9646845582e7665a1d5f91ace0c1e84ee58197fb4e6b91cb57c4c70367',
                },
            },
        },
    ]
    response = requests.post(
        'https://attend.gdcevents.gdconf.com/api/graphql', headers=headers_all, json=json_data_all)
    json_data = json.dumps(response.json(), indent=2)

    data_list = json.loads(json_data)
    list_id = []
    # Теперь переменная data_list содержит список словарей
    for item in data_list:
        # Предположим, что каждый словарь имеет ключ "pageInfo"
        # и этот ключ содержит информацию о странице, включая "endCursor"
        page_info = item["data"]["view"]["people"]["pageInfo"]
        end_cursor = page_info["endCursor"]

    for item in data_list:
        for node in item["data"]["view"]["people"]["nodes"]:
            list_id.append(node["id"])

    return end_cursor, list_id


def request_user_info(auth_key, personId):

    headers_user = {
        'accept': '*/*',
        'accept-language': 'ru,en;q=0.9',
        'authorization': auth_key,
        'content-type': 'application/json',
        # 'cookie': 'intercom-id-ineyz1by=e35f5b9f-b066-457a-9aff-0a53a0f551fb; intercom-device-id-ineyz1by=e53daf4f-460c-489c-ad59-9c84706e7f2f; swapcard-cookie-consent=%7B%22accepted%22%3Atrue%7D; __stripe_mid=e1417346-a801-431d-8686-ab32cbfd508a693eca; next-i18next=ru-RU; _ga=GA1.1.746820343.1711986307; swapcard-auth-api-refresh-token=eyJhbGciOiJSUzUxMiIsInR5cCI6IkpXVCJ9.eyJzZXNzaW9uSWQiOiI2NjBhZTI1YzAwMmZhNzBmNDI3MTFmOTciLCJ0eXBlIjoicmVmcmVzaC10b2tlbiIsInVzZXJJZCI6IjY0ZDYzZWMxMjZlMDlmNmE1NDI0ZGJmYiIsImlhdCI6MTcxMTk4OTM0MCwiaXNzIjoiYXV0aC1hcGkifQ.OEUEI9HJDC5FEa680wFow4abrnoVhfOnx57j7J-eMW7Iw8LQYOQB6Bqfew0ocPtaYR_AihztAG2cVYXl8IVbYkeqlMsf4DfkZpRe7Y1D5TRvtMqs-RkkTB6g-o0lAABwMZS5BA3M6pJ7ZjX9ol0F-sGSz6sgjBm-tq64sYGHUMBkg4v0KbhvZgQYRtGbtqC6CWmw5r043-pZ7jBTxrY_dNGjLSq03spnPisjgnqqH0-icib7XiiuNScqsemJhwVwI4uAdAMmHZCXk-7Y2B4_AKcSJOjp6YMiO-K1CkU-XQ1GeSQoCPjEbJGV1QbKqwQsugQvPUv2PLCumIx6OjGlOfJHhr383w2wlUNn7x-7GXvfNFR1u28I6ru0jJhmhYHcCsMbp6dtI6i-nUx8XRJv5Wz_EI8J9GzRcJ6Rkw1Y0uN9O89GDhzItQ3Jk8STEqCpFitwzOOvWDctrjnbXRsDVHj71OOxLq8VzLKVufO_U0ORvvD8zz3d3PayHLfLoe2NZEtxnRBJeJVjezDYHHDSHykMbml2J71WgCPRA52MRL9QGkwWrKMsQpWKyoOu9Nsj7IAVIelAJRLu7xrMBf04l9TH8edHRrHRJdPJprFt2N8qnNcjZbKlaLU-kQqEqjNSqzl8LLOg2xvvfU1y-hYQuN5BB2dwUB_cnKV4tdUqQYA; __stripe_sid=76346eba-8ea1-4bc0-bef5-621d4edc15437a7e56; intercom-session-ineyz1by=b3J1ekZXQmxTNHpjdUV6cG1hUTduenZSMWlLWmxIMDE3QkVCeVRzTHh6UklhcExpUjlKcUc2aEJCM3ZSVFhnRi0taXRwbjdMZjNpdDNrL2pEL3hVTmRvUT09--32abc8c0170998610a1d7d7a06f3fd43e3ae92a4; _ga_VWQ0TSBSYY=GS1.1.1712840847.12.1.1712844807.59.0.0',
        'origin': 'https://attend.gdcevents.gdconf.com',
        'referer': 'https://attend.gdcevents.gdconf.com/',
        'sec-ch-ua': '"Google Chrome";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
        'x-client-origin': 'attend.gdcevents.gdconf.com',
        'x-client-platform': 'Event App',
        'x-client-version': '2.308.37',
    }
    json_data_user = [
        {
            'operationName': 'EventPersonDetailsQuery',
            'variables': {
                'skipMeetings': True,
                'withEvent': True,
                'personId': personId,
                'userId': '',
                'eventId': 'RXZlbnRfMTY4MzkxNQ==',
            },
            'extensions': {
                'persistedQuery': {
                    'version': 1,
                    'sha256Hash': '8277c430fdaca5eb97969046e434f876a09af31e8e7940ed50ec234e9c052e77',
                },
            },
        },
    ]
    response = requests.post('https://attend.gdcevents.gdconf.com/api/graphql',
                             headers=headers_user, json=json_data_user)
    json_data = json.dumps(response.json(), indent=2)

    data_list = json.loads(json_data)
    # Теперь переменная data_list содержит список словарей
    try:
        workbook = load_workbook("output.xlsx")
    except FileNotFoundError:
        workbook = Workbook()

    # Выбираем активный лист или создаем новый, если лист не существует
    sheet = workbook.active

    # Заголовки столбцов
    headers = ["first_name", "last_name", "job_title",
               "email", "profile_link", "website_url", "phone"]

    # Проверяем, содержит ли лист данные и заголовки
    if sheet.max_row == 1 or [cell.value for cell in sheet[1]] != headers:
        # Если заголовки отсутствуют или не соответствуют ожидаемым, они добавляются
        sheet.insert_rows(1)
        for idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=idx, value=header)
    # Проходим по данным и добавляем их в таблицу
    # Множество для хранения уникальных комбинаций данных

    # Проход по данным и добавление их в таблицу
    for item in data_list:
        try:
            firstName = item["data"]["person"]["firstName"]
            lastName = item["data"]["person"]["lastName"]
            jobTitle = item["data"]["person"]["jobTitle"]
            email = item["data"]["person"]["email"]
            social_networks = item["data"]["person"]["socialNetworks"]
            linked_in = None  # Инициализация linked_in значением None, чтобы избежать ошибки, если данное поле отсутствует
            if email == None:
                break
            for network in social_networks:
                if network["type"] == "LINKEDIN":
                    linked_in = network["profile"] if network["profile"] else None
            websiteUrl = item["data"]["person"]["websiteUrl"] if item["data"]["person"]["websiteUrl"] else None
            landlinePhone = item["data"]["person"]["landlinePhone"][
                "formattedNumber"] if item["data"]["person"]["landlinePhone"] else None
            # Проверка и обновление linked_in, если он не начинается с "https://www.linkedin.com/in/"
            if linked_in and "https://" not in linked_in:
                linked_in = "https://www.linkedin.com/in/" + linked_in

            # Создание уникального идентификатора для текущей комбинации данных
            data_key = (firstName, lastName, jobTitle, email,
                        linked_in, websiteUrl, landlinePhone)

            # Проверка на повторение данных
            if email not in unique_data:
                unique_data.add(email)
                # Добавление данных в таблицу
                sheet.append([firstName, lastName, jobTitle, email,
                              linked_in, websiteUrl, landlinePhone])
                print("person added")

                # Добавление комбинации данных в множество уникальных данных

            else:
                break
            # Сохранение изменений в файле после каждого добавления

        except Exception as e:
            print(f"Ошибка при обработке данных: {e}")
        workbook.save("output.xlsx")


unique_data = set()
auth_key = request_login()
search = input("Введите ключевое слово для поиска: ")


def get_user_ids_and_end_cursor(auth_key, search, end_cursor=None):
    # Вызываем метод request_all с передачей текущего значения end_cursor
    end_cursor, list_id = request_all(
        auth_key, search=search, endCursor=end_cursor)
    # Возвращаем полученные значения
    return list_id, end_cursor


prev_end_cursor = None
end_cursor, list_id = request_all(auth_key, search=search)
while end_cursor != prev_end_cursor:
    # Сохраняем текущее значение end_cursor в prev_end_cursor
    prev_end_cursor = end_cursor
    # Получаем список идентификаторов пользователей и новое значение end_cursor
    list_id, end_cursor = get_user_ids_and_end_cursor(
        auth_key, search, end_cursor=end_cursor)
    # Перебираем идентификаторы пользователей и вызываем метод request_user_info для каждого идентификатора
    for id in list_id:
        request_user_info(auth_key, id)
print("parsing complete")

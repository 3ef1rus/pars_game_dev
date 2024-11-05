import json
from openpyxl import Workbook, load_workbook

with open("22.json", "r") as file:
    # Читаем содержимое файла
    json_data = file.read()

# Преобразуем содержимое JSON-файла в словарь
# Преобразуем JSON-строку в список словарей
data_list = json.loads(json_data)

# Создаем новую книгу Excel или загружаем существующую, если файл уже существует
try:
    workbook = load_workbook("output.xlsx")
except FileNotFoundError:
    workbook = Workbook()

# Выбираем активный лист или создаем новый, если лист не существует
sheet = workbook.active

# Заголовки столбцов
headers = ["first_name", "last_name", "job_title",
           "email", "profile_link", "website_url", "phone"]

# Проверяем, содержит ли лист данные и заголовки
if sheet.max_row == 1 or [cell.value for cell in sheet[1]] != headers:
    # Если заголовки отсутствуют или не соответствуют ожидаемым, они добавляются
    sheet.append(headers)

# Проходим по данным и добавляем их в таблицу
for item in data_list:
    try:
        firstName = item["data"]["person"]["firstName"]
        lastName = item["data"]["person"]["lastName"]
        jobTitle = item["data"]["person"]["jobTitle"]
        email = item["data"]["person"]["email"]
        social_networks = item["data"]["person"]["socialNetworks"]
        linked_in = None  # Инициализируем linked_in значением None, чтобы избежать ошибки, если данное поле отсутствует
        for network in social_networks:
            if network["type"] == "LINKEDIN":
                linked_in = network["profile"]
        websiteUrl = item["data"]["person"]["websiteUrl"]
        landlinePhone = item["data"]["person"]["landlinePhone"]["formattedNumber"]
        if "https://www.linkedin.com/in/" not in linked_in:
            linked_in = "https://www.linkedin.com/in/"+linked_in
            # Добавляем данные в таблицу
        sheet.append([firstName, lastName, jobTitle, email,
                      linked_in, websiteUrl, landlinePhone])
        print("person added")

    except Exception as e:
        pass

# Сохраняем изменения в файл
workbook.save("output.xlsx")
